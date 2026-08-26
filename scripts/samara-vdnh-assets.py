#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Ассеты кейса «Стенд Самарской области на ВДНХ» (/portfolio/samara-stand-vdnh).

Источник: ~/Documents/Материалы для обновления сайта/Самара ВДНХ (плюс подпапки
«Монтаж», «Фирменный стиль», «Эмуляции...», «Предвариетльные рендеры...»)
и шесть файлов «Календарь*.xlsx» из ~/Downloads.

Что делает:
  1. naked eye. Из презентации «ОПРЕДЕЛЕНИЕ ТОЧКИ РЕНДЕРА» режет пары рендеров
     (общий план стенда и вид с точки зрителя) для матрицы «камера × обзор»,
     плюс план павильона с выбранной точкой.
  2. Раскадровка паруса: шесть сцен нарезаются из одного листа сценарного плана.
  3. Кинетический экран: постеры четырёх эмуляций и веб-версии самих роликов.
  4. Фотографии: монтаж и работа стенда, ресайз до 1600 px и webp.
  5. Расписание: шесть недельных календарей сводятся в schedule.json
     (слоты, форматы, тематические недели).

Запуск: python3 scripts/samara-vdnh-assets.py [--skip-video]
Идемпотентен, готовые файлы перезаписывает.
"""
import json
import os
import re
import subprocess
import sys
import datetime
from collections import Counter

import cv2
import numpy as np
import fitz
from PIL import Image, ImageOps

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, '..'))
SRC = os.path.expanduser('~/Documents/Материалы для обновления сайта/Самара ВДНХ')
MONT = os.path.join(SRC, 'Монтаж')
EMUL = os.path.join(SRC, 'Эмуляции выборочных проектов для кинетического  экрана')
ROCKET = os.path.join(SRC, 'Предвариетльные рендеры на согласоание запуска ракеты')
STYLE = os.path.join(SRC, 'Фирменный стиль')
DOWNLOADS = os.path.expanduser('~/Downloads')

OUT = os.path.join(ROOT, 'mirror', 'images', 'lib', 'custom-samara-vdnh')
MEDIA = os.path.join(ROOT, 'mirror', 'media')
DATA = os.path.join(ROOT, 'scripts', 'a2', 'samara_vdnh_map.json')

NK_PDF = os.path.join(SRC, 'ОПРЕДЕЛЕНИЕ ТОЧКИ РЕНДЕРА ДЛЯ nided eye.pdf')

# ─── матрица naked eye ──────────────────────────────────────────────────────
# страница PDF (1-based) → (камера, точка зрителя, сюжет на экране)
NK_PAGES = [
    (4,  'right',  'right',  'ladya'),
    (5,  'right',  'left',   'ladya'),
    (6,  'right',  'center', 'ladya'),
    (8,  'center', 'right',  'ladya'),
    (9,  'center', 'center', 'ladya'),
    (10, 'center', 'left',   'ladya'),
    (11, 'center', 'right',  'rocket'),
    (12, 'center', 'center', 'rocket'),
    (13, 'center', 'left',   'rocket'),
]

# ─── фотографии: (файл, имя на выходе, подпись) ─────────────────────────────
PHOTOS_MONTAGE = [
    ('IMG_050995C147EB-1.jpeg', 'mnt-frame',   'Каркас паруса: ламели набираются на дугу до того, как встанет экран'),
    ('IMG_489ACB1228B8-1.jpeg', 'mnt-sphere',  'LED-шар собирают в кольце, панель за панелью'),
    ('IMG_F6883ADB6BE8-1.jpeg', 'mnt-guts',    'Внутренности шара: развязка питания и сигнала'),
    ('IMG_FE5C1CEF9F4B-1.jpeg', 'mnt-modules', 'Модули экрана ставят по одному, с выверкой по ряду'),
    ('IMG_B2071EE99A88-1.jpeg', 'mnt-arc',     'Дуга экрана вблизи: обратная сторона паруса'),
    ('IMG_FF45F2669196-1.jpeg', 'mnt-cables',  'Кабельные трассы над стендом: 400 метров сигнала и 400 силовых'),
    ('IMG_F07D1BA59021-1.jpeg', 'mnt-render',  'Сборка контента для кинетического экрана в работе'),
    ('IMG_9B6ED65B25A5-1.jpeg', 'mnt-globe',   'Шар включили: первая проверка картинки на площадке'),
]
PHOTOS_LIVE = [
    ('20240124-DSC04325.jpg',   'live-vr',      'VR-кинотеатр: фильм о регионе в очках'),
    ('20240124-DSC04292.jpg',   'live-rocket',  'Ракета на парусе и зрители перед экраном'),
    ('20240124-DSC03976.jpg',   'live-stage',   'Выступление на фоне паруса'),
    ('20240124-DSC04123.jpg',   'live-choir',   'Творческие коллективы региона на стенде'),
    ('IMG_AC1AAB9A82AF-1.jpeg', 'live-kinect',  'Kinect-футбол: ворота на экране, вратарь перед ним'),
    ('IMG_D98D6DAFB036-1.jpeg', 'live-touch',   'Данные региона на большой панели'),
    ('IMG_082618685967-1.jpeg', 'live-panels',  'Вертикальные тач-панели по периметру стенда'),
    ('IMG_C6A13C557D65-1.jpeg', 'live-kokosh',  'Костюмы с самарскими сюжетами'),
    ('IMG_B52FFE3258A1-1.jpeg', 'live-gov',     'Официальная часть на стенде'),
    ('IMG_1E61E2B98673-1.jpeg', 'live-map',     'Карта области на экране и участники программы'),
    ('20240412_105658.jpg',     'live-lecture', 'Дневная программа: зал перед экраном заполнен'),
    ('IMG_ACED77B017FB-1.jpeg', 'live-family',  'Семейные съемки на фоне стенда'),
]
PHOTO_BUILD = ('застройка.jpg', 'build-hall', 'Павильон за несколько недель до открытия')
PHOTO_TEAM = ('Часть нашей команды.jpeg', 'team', 'Часть команды проекта на площадке')

# ─── эмуляции кинетического экрана ──────────────────────────────────────────
EMULATIONS = [
    ('Самара_в_лицах_1.mp4',      'kin-faces',  'Самара в лицах',
     'Портреты земляков во весь экран: рельеф собирает лицо, подсветка добавляет объем.'),
    ('Жены героев_1.mp4',         'kin-wives',  'Жены героев',
     'Портретная серия с текстовыми блоками, самая длинная из программ.'),
    ('Спорт в лицах_0_INTRO-2.mp4', 'kin-sport', 'Спорт в лицах',
     'Заставка спортивного блока: движение по рядам пикселей вместо монтажной склейки.'),
    ('было_стало_1.mp4',          'kin-before', 'Было и стало',
     'Сравнение городских объектов: рельеф переключает состояние прямо на плоскости экрана.'),
]

STORY_SHEET = os.path.join(SRC, 'Сценарный план для 3д контента на жкран парус.jpg')
# названия сцен по порядку сверху вниз; границы блоков ищем по чёрным полосам листа
STORY_TITLES = [
    'Природа Самарской области',
    'Образ ладьи',
    'Мост через Волгу',
    'Площадь Славы',
    'Ракета «Союз»',
    '53,6 тысячи километров',
]


def sh(cmd):
    subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


SIZES = {}          # имя кадра → (ширина, высота) готового файла


def save(img, name, width=1600, quality=84):
    """PIL/np → jpg + webp в OUT. Разворачивает кадр по EXIF."""
    if isinstance(img, np.ndarray):
        img = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
    else:
        img = ImageOps.exif_transpose(img)      # снимки с телефона лежат боком
    if img.mode not in ('RGB',):
        img = img.convert('RGB')
    if img.width > width:
        img = img.resize((width, round(img.height * width / img.width)), Image.LANCZOS)
    p = os.path.join(OUT, name + '.jpg')
    img.save(p, quality=quality, optimize=True)
    img.save(p + '.webp', quality=78, method=5)
    SIZES[name] = [img.width, img.height]
    return os.path.getsize(p)


def boxes_on_page(png):
    """Прямоугольники картинок на отрендеренной полосе (подписи внизу отрезаны)."""
    im = cv2.imread(png)
    H, W = im.shape[:2]
    g = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    e = cv2.Canny(cv2.GaussianBlur(g, (5, 5), 0), 25, 80)
    e[int(H * .86):, :] = 0
    e = cv2.dilate(e, np.ones((25, 25), np.uint8), 1)
    n, _lab, stats, _c = cv2.connectedComponentsWithStats(e, 8)
    out = []
    for i in range(1, n):
        x, y, w, h, _a = stats[i]
        if w * h < .02 * W * H:
            continue
        out.append([int(x), int(y), int(w), int(h)])
    out.sort(key=lambda b: b[0])
    return im, out


def crop(im, box, pad=24):
    H, W = im.shape[:2]
    x, y, w, h = box
    x0, y0 = max(0, x - pad), max(0, y - pad)
    x1, y1 = min(W, x + w + pad), min(H, y + h + pad)
    return im[y0:y1, x0:x1]


def naked_eye(man):
    """Пары рендеров матрицы и план павильона."""
    doc = fitz.open(NK_PDF)
    tmp = os.path.join(OUT, '_tmp.png')
    grid = []
    for page, cam, view, scene in NK_PAGES:
        doc[page - 1].get_pixmap(matrix=fitz.Matrix(2, 2)).save(tmp)
        im, bx = boxes_on_page(tmp)
        base = 'nk-%s-%s-%s' % (scene, cam, view)
        only_wide = len(bx) < 2
        if only_wide:
            # на полосе сохранился только общий план: крупного вида для этой
            # пары в согласовании не рисовали, подписываем это на странице
            wide = close = bx[0]
        else:
            wide, close = bx[0], bx[-1]
        save(crop(im, close), base, width=1100)
        save(crop(im, wide), base + '-wide', width=1100)
        grid.append({'scene': scene, 'camera': cam, 'view': view,
                     'close': base, 'wide': base + '-wide', 'only_wide': only_wide})
    # план павильона с выбранной точкой (полоса «Вывод»)
    doc[14].get_pixmap(matrix=fitz.Matrix(2, 2)).save(tmp)
    im, bx = boxes_on_page(tmp)
    if bx:
        save(crop(im, max(bx, key=lambda b: b[2] * b[3]), pad=40), 'nk-plan', width=1200)
    os.remove(tmp)
    man['nakedeye'] = grid
    print('  naked eye: %d рендеров + план' % (len(grid) * 2))


def storyboard(man):
    """Лист сценарного плана режется по чёрным промежуткам между сценами."""
    im = Image.open(STORY_SHEET)
    a = np.asarray(im.convert('L'), dtype='float32')
    H = a.shape[0]
    light = (a > 60).mean(axis=1)
    gaps, run = [], None
    for y, v in enumerate(light):
        if v < .02:
            run = y if run is None else run
        elif run is not None:
            if y - run >= 14:
                gaps.append((run, y))
            run = None
    if run is not None:
        gaps.append((run, H))
    # границы сцен = промежутки между полосами пустоты
    bounds, prev = [], 0
    for g0, g1 in gaps:
        if g0 - prev > H * .06:
            bounds.append((prev, g0))
        prev = g1
    if H - prev > H * .06:
        bounds.append((prev, H))
    rows = []
    for i, title in enumerate(STORY_TITLES):
        if i >= len(bounds):
            break
        y0, y1 = bounds[i]
        name = 'story-%d' % (i + 1)
        save(im.crop((0, y0, im.width, y1)), name, width=1400)
        rows.append({'title': title, 'img': name})
    man['storyboard'] = rows
    print('  раскадровка: %d сцен, границы по листу' % len(rows))


def photos(man):
    out = {'montage': [], 'live': []}
    for src_dir, items, key in ((MONT, PHOTOS_MONTAGE, 'montage'),
                                (SRC, PHOTOS_LIVE, 'live')):
        for fn, name, cap in items:
            p = os.path.join(src_dir, fn)
            if not os.path.exists(p):
                print('  ! нет файла', fn)
                continue
            save(Image.open(p), name, width=1600)
            out[key].append({'img': name, 'cap': cap})
    for fn, name, cap in (PHOTO_BUILD, PHOTO_TEAM):
        p = os.path.join(SRC, fn)
        if os.path.exists(p):
            save(Image.open(p), name, width=1600)
            out[name] = {'img': name, 'cap': cap}
    man['photos'] = out
    print('  фото: монтаж %d, работа стенда %d' % (len(out['montage']), len(out['live'])))


def kinetic(man, skip_video=False):
    items = []
    for fn, name, title, note in EMULATIONS:
        p = os.path.join(EMUL, fn)
        if not os.path.exists(p):
            print('  ! нет эмуляции', fn)
            continue
        dur = float(subprocess.run(
            ['ffprobe', '-v', 'error', '-select_streams', 'v', '-show_entries',
             'stream=duration', '-of', 'csv=p=0', p],
            capture_output=True, text=True).stdout.strip().rstrip(','))
        poster = os.path.join(OUT, name + '.jpg')
        sh(['ffmpeg', '-y', '-loglevel', 'error', '-ss', str(min(12, dur / 3)), '-i', p,
            '-frames:v', '1', '-vf', 'scale=1400:-2', poster])
        Image.open(poster).save(poster + '.webp', quality=78, method=5)
        web = os.path.join(MEDIA, 'samara-vdnh-%s.mp4' % name.replace('kin-', 'kinetic-'))
        if not skip_video:
            sh(['ffmpeg', '-y', '-loglevel', 'error', '-i', p, '-vf', 'scale=960:-2',
                '-c:v', 'libx264', '-crf', '28', '-preset', 'slow', '-pix_fmt', 'yuv420p',
                '-movflags', '+faststart', '-an', web])
        items.append({'name': name, 'title': title, 'note': note,
                      'dur': round(dur), 'video': '/media/' + os.path.basename(web),
                      'mb': round(os.path.getsize(web) / 1e6, 1) if os.path.exists(web) else None})
    man['kinetic'] = items
    print('  кинетика: %d программ' % len(items))


def parus(man, skip_video=False):
    """Ролик для экрана-паруса в родной вертикальной пропорции."""
    p = os.path.join(SRC, 'sc_parus_v05.mp4')
    vids = os.path.join(ROOT, 'mirror', 'videos')
    os.makedirs(vids, exist_ok=True)
    web = os.path.join(vids, 'samara-vdnh-parus.mp4')
    poster = os.path.join(OUT, 'parus-poster.jpg')
    if not os.path.exists(p):
        return
    if not skip_video:
        sh(['ffmpeg', '-y', '-loglevel', 'error', '-i', p, '-vf', 'scale=900:-2',
            '-c:v', 'libx264', '-crf', '24', '-preset', 'slow', '-pix_fmt', 'yuv420p',
            '-movflags', '+faststart', '-an', web])
    sh(['ffmpeg', '-y', '-loglevel', 'error', '-ss', '6', '-i', p, '-frames:v', '1',
        '-vf', 'scale=900:-2', poster])
    Image.open(poster).save(poster + '.webp', quality=78, method=5)
    man['parus'] = {'video': '/videos/samara-vdnh-parus.mp4', 'poster': 'parus-poster',
                    'mb': round(os.path.getsize(web) / 1e6, 1) if os.path.exists(web) else None}
    print('  парус: ролик и постер готовы')


def schedule(man):
    """Шесть недельных календарей → слоты, форматы, тематические недели."""
    import openpyxl
    slots = []
    weeks = []
    for fn in sorted(f for f in os.listdir(DOWNLOADS) if f.startswith('Календарь')
                     and f.endswith('.xlsx')):
        wb = openpyxl.load_workbook(os.path.join(DOWNLOADS, fn), data_only=True)
        ws = wb['Календарь'] if 'Календарь' in wb.sheetnames else wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # тема недели надёжнее всего лежит в имени файла
        # имена файлов с macOS приходят в NFD: «й» это «и» плюс бреве
        import unicodedata
        title = unicodedata.normalize('NFC', os.path.splitext(fn)[0])
        title = re.sub(r'^Календарь\.?\s*', '', title).strip()
        title = re.sub(r'\s+', ' ', title.replace('.', ',')).strip(' ,')
        cnt = 0
        for r in rows:
            for i, c in enumerate(r):
                if isinstance(c, datetime.time) and i + 1 < len(r):
                    nxt = r[i + 1]
                    if nxt and str(nxt).strip():
                        name = re.sub(r'"([^"]+)"', '«\\1»', str(nxt).strip())
                        slots.append({'t': c.strftime('%H:%M'), 'e': name})
                        cnt += 1
        weeks.append({'file': fn, 'title': title, 'slots': cnt})
    kinds = Counter(s['e'] for s in slots)
    hours = Counter(s['t'] for s in slots)
    man['schedule'] = {
        'total': len(slots),
        'unique': len(kinds),
        'weeks': weeks,
        'top': kinds.most_common(28),
        'byhour': sorted(hours.items()),
    }
    print('  расписание: %d слотов, %d форматов' % (len(slots), len(kinds)))



# ─── шрифты: Tenor Sans (заголовки) + Source Sans 3 (текст) + JetBrains Mono ─
UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/124.0 Safari/537.36')
GF = ('https://fonts.googleapis.com/css2?family=Tenor+Sans'
      '&family=Source+Sans+3:wght@300;400;600;700'
      '&family=JetBrains+Mono:wght@400;600&display=swap')
KEEP_SUBSETS = ('cyrillic-ext', 'cyrillic', 'latin-ext', 'latin')
FONTS_DIR = os.path.join(ROOT, 'mirror', 'fonts')
FONTS_FILES = os.path.join(FONTS_DIR, 'files')


def fetch(url):
    import urllib.request
    req = urllib.request.Request(url, headers={'User-Agent': UA})
    return urllib.request.urlopen(req).read()


def fonts():
    """Кладёт шрифты локально: внешних CDN на сайте нет принципиально."""
    os.makedirs(FONTS_FILES, exist_ok=True)
    css = fetch(GF).decode('utf-8')
    blocks = re.findall(r'/\*\s*([a-z-]+)\s*\*/\s*(@font-face\s*\{.*?\})', css, re.S)
    out, n = [], 0
    for subset, block in blocks:
        if subset not in KEEP_SUBSETS:
            continue
        fam = re.search(r"font-family:\s*'([^']+)'", block).group(1)
        wght = re.search(r'font-weight:\s*(\d+)', block).group(1)
        name = '%s-%s-%s.woff2' % (fam.lower().replace(' ', '-'), wght, subset)
        url = re.search(r'url\((https://[^)]+)\)', block).group(1)
        path = os.path.join(FONTS_FILES, name)
        if not os.path.exists(path):
            open(path, 'wb').write(fetch(url))
            n += 1
        out.append(block.replace(url, 'files/' + name))
    head = ('/* Tenor Sans + Source Sans 3 + JetBrains Mono, self-host для\n'
            '   /portfolio/samara-stand-vdnh/. Сгенерировано\n'
            '   scripts/samara-vdnh-assets.py, руками не править. */\n')
    open(os.path.join(FONTS_DIR, 'tenor-source.css'), 'w', encoding='utf-8').write(
        head + '\n'.join(out) + '\n')
    print('  шрифты: %d @font-face, скачано %d' % (len(out), n))



# ─── карты рельефа для живой сетки кинетического экрана ─────────────────────
# 64×36 значений 0..1: столько «пикселей» отрисовывает страница. Источники
# честные: знак выставки из фирстиля, герб с ролика паруса, кадр эмуляции.
MASK_W, MASK_H = 104, 58


def _mask_from(arr):
    """np.float32 (H×W, 0..1) → сетка MASK_H×MASK_W, растянутая по контрасту."""
    im = Image.fromarray((np.clip(arr, 0, 1) * 255).astype('uint8'))
    im = im.resize((MASK_W, MASK_H), Image.LANCZOS)
    a = np.asarray(im, dtype='float32') / 255.0
    lo, hi = float(np.percentile(a, 2)), float(np.percentile(a, 98))
    if hi - lo > 1e-3:
        a = np.clip((a - lo) / (hi - lo), 0, 1)
    a = a ** 0.78                      # подтягиваем середину: рельеф читается лучше
    return [[round(float(v), 3) for v in row] for row in a]


def kinetic_masks(man):
    out = {}
    # 1. знак выставки: берём альфу логотипа, вписываем в 16:9
    logo = os.path.join(STYLE, 'пнг лого', 'лого ВДНХ без фона-01.png')
    if os.path.exists(logo):
        a = np.asarray(Image.open(logo).convert('RGBA'), dtype='float32')[:, :, 3] / 255.0
        ys, xs = np.where(a > 0.08)
        a = a[ys.min():ys.max() + 1, xs.min():xs.max() + 1]
        h, w = a.shape
        ch = int(h * 1.16)                      # поля вокруг знака
        cw = max(int(ch * 16 / 9), int(w * 1.16))
        ch = max(ch, int(cw * 9 / 16))
        canvas = np.zeros((ch, cw), dtype='float32')
        y0, x0 = (ch - h) // 2, (cw - w) // 2
        canvas[y0:y0 + h, x0:x0 + w] = a
        out['ladya'] = _mask_from(canvas)
    # 2. герб Самарской области с ролика для паруса
    poster = os.path.join(OUT, 'parus-poster.jpg')
    if os.path.exists(poster):
        g = np.asarray(Image.open(poster).convert('L'), dtype='float32') / 255.0
        h, w = g.shape
        g = g[int(h * .30):int(h * .62), int(w * .30):int(w * .70)]
        out['gerb'] = _mask_from(1.0 - g)      # герб темнее фона, инвертируем
    # 3. портрет из эмуляции «Самара в лицах»
    src = os.path.join(EMUL, 'Самара_в_лицах_1.mp4')
    if os.path.exists(src):
        tmp = os.path.join(OUT, '_face.png')
        sh(['ffmpeg', '-y', '-loglevel', 'error', '-ss', '8', '-i', src,
            '-frames:v', '1', tmp])
        g = np.asarray(Image.open(tmp).convert('L'), dtype='float32') / 255.0
        h, w = g.shape
        g = g[int(h * .06):int(h * .97), int(w * .13):int(w * .67)]
        out['face'] = _mask_from(g)
        os.remove(tmp)
    man['masks'] = {'w': MASK_W, 'h': MASK_H, 'maps': out}
    print('  рельеф: %d карт %d×%d' % (len(out), MASK_W, MASK_H))



# ─── кадры из роликов проекта: общий план стенда и живой кинетический экран ─
# секунды выверены по контактному листу (fps=1/6 по samara-vdnh-1.mp4)
STAND_FRAMES = [
    ('mirror/media/samara-vdnh-1.mp4', 42.0, 'stand-hero',
     'Стенд Самарской области в павильоне 75: изогнутый парус, амфитеатр, витрины'),
    ('mirror/media/samara-vdnh-1.mp4', 36.5, 'stand-front',
     'Ладья снаружи: парус работает обложкой стенда'),
    ('mirror/media/samara-vdnh-1.mp4', 14.0, 'stand-gerb',
     'Объемный герб области во весь экран-парус'),
    ('mirror/media/samara-vdnh-1.mp4', 150.5, 'stand-balls',
     'Три LED-шара над стендом: Земля, Луна и Марс'),
    ('mirror/media/samara-vdnh-1.mp4', 22.0, 'stand-amphi',
     'Амфитеатр внутри корпуса ладьи: здесь шла живая программа'),
    ('mirror/media/samara-vdnh-1.mp4', 128.0, 'kin-live',
     'Кинетический экран сбоку: пиксели действительно выдвинуты из плоскости'),
    ('mirror/media/samara-vdnh-1.mp4', 126.5, 'kin-live2',
     'Портрет на рельефе: проект «Самарский спорт в лицах»'),
    ('mirror/media/samara-vdnh-1.mp4', 20.0, 'kin-text',
     'Рукописный текст на рельефе: письмо с фронта в проекте «Лица армии»'),
    ('mirror/media/samara-vdnh-1.mp4', 116.0, 'kin-logo',
     'Знак выставки и области, собранный выдвинутыми пикселями'),
    ('mirror/media/samara-vdnh-1.mp4', 96.0, 'stand-touch',
     'Интерфейс тач-панели: игры и разделы о регионе'),
    ('mirror/media/samara-vdnh-1.mp4', 110.0, 'stand-basket',
     'Kinect-баскетбол: бросок рукой по экрану'),
    ('mirror/media/samara-vdnh-1.mp4', 18.0, 'stand-transparent',
     'Прозрачные панели: текст ложится поверх витрины с айдентикой'),
    ('mirror/media/samara-vdnh-1.mp4', 166.0, 'stand-rocket',
     'Ракета «Союз» и костюмы с самарскими сюжетами'),
    ('mirror/media/samara-vdnh-1.mp4', 172.0, 'stand-crowd',
     'Открытие: у стенда собралась толпа'),
]


def stand_frames(man):
    out = []
    for src, sec, name, cap in STAND_FRAMES:
        path = os.path.join(ROOT, src)
        if not os.path.exists(path):
            print('  ! нет ролика', src)
            continue
        tmp = os.path.join(OUT, '_frame.png')
        sh(['ffmpeg', '-y', '-loglevel', 'error', '-ss', str(sec), '-i', path,
            '-frames:v', '1', tmp])
        save(Image.open(tmp), name, width=1600, quality=86)
        os.remove(tmp)
        out.append({'img': name, 'cap': cap})
    man['frames'] = out
    print('  кадры со стенда: %d' % len(out))



# ─── вся фотобаза проекта ───────────────────────────────────────────────────
# В галерею идут снимки со стенда и с монтажа. Служебные листы (раскадровки,
# рендеры, фирстиль) отсеиваем по именам: они живут в своих блоках.
GAL_SKIP = (
    'Сценарный план', 'Начальная раскодровка', 'начальная раскадровка',
    'предварительный рендер', 'создания контнта', 'Изображение JPEG',
    'telegram-cloud-photo', 'Часть нашей команды', 'застройка',
)
GAL_DIR = 'g'


def _gal_ok(name):
    return not any(k.lower() in name.lower() for k in GAL_SKIP)


def _dhash(im, size=10):
    """Перцептивный хэш: разница соседних пикселей на сетке size×(size+1)."""
    g = im.convert('L').resize((size + 1, size), Image.LANCZOS)
    a = np.asarray(g, dtype='int16')
    bits = (a[:, 1:] > a[:, :-1]).flatten()
    v = 0
    for b in bits:
        v = (v << 1) | int(b)
    return v


def gallery_all(man):
    """Все кадры проекта: 1200 px, jpg + webp. Дубли снимаем по dhash."""
    import hashlib
    out_dir = os.path.join(OUT, GAL_DIR)
    os.makedirs(out_dir, exist_ok=True)
    files = []
    for src_dir, group in ((SRC, 'stand'), (MONT, 'montage')):
        for fn in sorted(os.listdir(src_dir)):
            if not fn.lower().endswith(('.jpg', '.jpeg')) or not _gal_ok(fn):
                continue
            files.append((os.path.join(src_dir, fn), group, fn))
    items, hashes, dups = [], [], 0
    for path, group, fn in files:
        try:
            im = ImageOps.exif_transpose(Image.open(path)).convert('RGB')
        except Exception:
            continue
        w, h = im.size
        if max(w, h) < 900:                     # мелкие превью в галерею не берём
            continue
        # почти одинаковые кадры (серии из двух-трёх снимков подряд) оставляем в одном
        hv = _dhash(im)
        if any(bin(hv ^ prev).count('1') <= 8 for prev in hashes):
            dups += 1
            continue
        hashes.append(hv)
        name = hashlib.md5(fn.encode('utf-8')).hexdigest()[:10]
        if im.width > 1200:
            im = im.resize((1200, round(h * 1200 / w)), Image.LANCZOS)
        p = os.path.join(out_dir, name + '.jpg')
        im.save(p, quality=74, optimize=True)
        im.save(p + '.webp', quality=68, method=5)
        items.append({'n': name, 'g': group, 'r': round(im.height / im.width, 3)})
    man['gallery_all'] = items
    keep = {i['n'] + ext for i in items for ext in ('.jpg', '.jpg.webp')}
    for f in os.listdir(out_dir):               # чистим файлы отсеянных дублей
        if f not in keep:
            os.remove(os.path.join(out_dir, f))
    size = sum(os.path.getsize(os.path.join(out_dir, f)) for f in os.listdir(out_dir))
    print('  фотобаза: %d кадров (дублей убрано %d), %.1f МБ' % (len(items), dups, size / 1e6))



# ─── ТЗ на кинетический экран (3D Transformer Screen, ПМТ) ──────────────────
TZ_PDF = os.path.expanduser('~/Downloads/ТЗ_для_застройщика_Самарская_область_v2_1.pdf')
# страница PDF (1-based) → (имя, доля обрезки сверху/снизу, подпись)
TZ_PAGES = [
    (3, 'tz-cabinet', (.28, .78), 'Один кабинет: 644 × 483 мм, глубина 320, вес 25 кг'),
    (4, 'tz-wall', (.02, .34), 'Стена 5 × 4: двадцать кабинетов, 3220 × 1932 мм, 500 кг'),
    (9, 'tz-plan', (.10, .70), 'Расположение решений на стенде: кинетическая стена и серверная'),
]


def tz_frames(man):
    doc = fitz.open(TZ_PDF)
    out = []
    for page, name, (y0, y1), cap in TZ_PAGES:
        pix = doc[page - 1].get_pixmap(matrix=fitz.Matrix(2.4, 2.4))
        tmp = os.path.join(OUT, '_tz.png')
        pix.save(tmp)
        im = Image.open(tmp)
        im = im.crop((0, int(im.height * y0), im.width, int(im.height * y1)))
        # обрезаем белые поля листа, чтобы чертёж занимал всю ширину
        a = np.asarray(im.convert('L'), dtype='uint8')
        ys, xs = np.where(a < 235)
        if len(xs):
            pad = 18
            im = im.crop((max(0, xs.min() - pad), max(0, ys.min() - pad),
                          min(im.width, xs.max() + pad), min(im.height, ys.max() + pad)))
        save(im, name, width=1500, quality=88)
        os.remove(tmp)
        out.append({'img': name, 'cap': cap})
    man['tz'] = out
    print('  схемы из ТЗ: %d' % len(out))



# ─── полосы клиентской концепции (Концепция_Самара_v5) ─────────────────────
CONCEPT_PDF = os.path.join(SRC, 'Концепция_Самара_v5_230914.pdf')
CONCEPT_PAGES = [
    (20, 'cn-visual', 'Визуальный язык проекта: движение, динамика, изменения, открытия'),
    (21, 'cn-mood1', 'Референсы визуального ряда из концепции'),
    (24, 'cn-mood2', 'Референсы визуального ряда из концепции'),
    (27, 'cn-sail', 'Фронтальный экран-парус: описание из концепции'),
    (29, 'cn-kinetic', 'Кинетический экран в концепции'),
    (30, 'cn-touch', 'Восемь тач-панелей: размещение и режим ожидания'),
    (45, 'cn-menu', 'Структура меню тач-панели из концепции'),
]


def concept_frames(man):
    doc = fitz.open(CONCEPT_PDF)
    out = []
    for page, name, cap in CONCEPT_PAGES:
        pix = doc[page - 1].get_pixmap(matrix=fitz.Matrix(2, 2))
        tmp = os.path.join(OUT, '_cn.png')
        pix.save(tmp)
        save(Image.open(tmp), name, width=1500, quality=86)
        os.remove(tmp)
        out.append({'img': name, 'cap': cap})
    man['concept'] = out
    print('  полосы концепции: %d' % len(out))



# ─── что показывает модель кинетического экрана ─────────────────────────────
# Реальная стена: 5 × 4 кабинета, в кабинете 4 × 3 подвижных блока,
# то есть 20 × 12 = 240 блоков на всю стену. Пропорция 3220 × 1932 мм.
KIN_GRID = (20, 12)
KIN_SOURCES = [
    ('face', 'Портрет', 'kin-live2', (.36, .04, .99, .93)),
    ('gerb', 'Герб области', 'stand-gerb', (.30, .12, .70, .92)),
    ('stand', 'Кадр со стенда', 'stand-front', (.10, .05, .92, .95)),
]


def kinetic_sources(man):
    """Плоские кадры 5:3 под модель: их страница режет на 20 × 12 блоков."""
    out = []
    W, H = 1000, 600
    def put(name, im, title):
        im = im.convert('RGB')
        # кадрируем по центру в пропорцию стены
        r = W / H
        if im.width / im.height > r:
            nw = int(im.height * r)
            im = im.crop(((im.width - nw) // 2, 0, (im.width + nw) // 2, im.height))
        else:
            nh = int(im.width / r)
            im = im.crop((0, (im.height - nh) // 2, im.width, (im.height + nh) // 2))
        im = im.resize((W, H), Image.LANCZOS)
        save(im, 'kinsrc-' + name, width=W, quality=86)
        out.append({'id': name, 'title': title, 'img': 'kinsrc-' + name})

    for name, title, src, box in KIN_SOURCES:
        if name == 'city':   # осталось на случай возврата кадра из эмуляции
            # кадр из эмуляции «было и стало»: там площадка снята фронтально
            p = os.path.join(EMUL, 'было_стало_1.mp4')
            if not os.path.exists(p):
                continue
            tmp = os.path.join(OUT, '_city.png')
            sh(['ffmpeg', '-y', '-loglevel', 'error', '-ss', '96', '-i', p,
                '-frames:v', '1', tmp])
            im = Image.open(tmp)
            im = im.crop((int(im.width * .17), int(im.height * .12),
                          int(im.width * .83), int(im.height * .80)))
            put(name, im, title)
            os.remove(tmp)
            continue
        p = os.path.join(OUT, src + '.jpg')
        if not os.path.exists(p):
            continue
        im = Image.open(p)
        if box:
            im = im.crop((int(im.width * box[0]), int(im.height * box[1]),
                          int(im.width * box[2]), int(im.height * box[3])))
        put(name, im, title)
    man['kinsrc'] = {'grid': list(KIN_GRID), 'items': out}
    print('  источники для модели экрана: %d' % len(out))


def main():
    skip_video = '--skip-video' in sys.argv
    os.makedirs(OUT, exist_ok=True)
    man = {}
    print('Самара ВДНХ: ассеты')
    fonts()
    naked_eye(man)
    storyboard(man)
    photos(man)
    kinetic(man, skip_video)
    kinetic_masks(man)
    stand_frames(man)
    kinetic_sources(man)
    tz_frames(man)
    concept_frames(man)
    gallery_all(man)
    parus(man, skip_video)
    schedule(man)
    man['sizes'] = SIZES
    with open(DATA, 'w', encoding='utf-8') as f:
        json.dump(man, f, ensure_ascii=False, indent=1)
    size = sum(os.path.getsize(os.path.join(OUT, f)) for f in os.listdir(OUT))
    print('готово: %s, %.1f МБ' % (os.path.relpath(OUT, ROOT), size / 1e6))


if __name__ == '__main__':
    main()
